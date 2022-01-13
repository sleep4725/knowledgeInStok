from builtins import filter

import xlsxwriter
from openpyxl import Workbook
from openpyxl.workbook import workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.styles import numbers, Alignment, Font, PatternFill

## ==============================
# 작성자 : 김준현
# 작성일 : 2022-01-13
## ==============================
class XlClient:

    def __init__(self):
        # 엑셀파일 쓰기
        self.write_wb = Workbook()
        self.write_wb.guess_type = False
        self.reporting_file_path = "./ResultReporting/{_file_name_}.xlsx"
        self.xl_column = ["num", "name", "code", "PER", "PBR", "ROE", "EPS", "BPS", "5년ROE", "5년EPS성장률"]

        self.write_wb.create_sheet('주식정보')
        self.write_ws = self.write_wb["주식정보"]

        sheet1 = self.write_wb["Sheet"]
        self.write_wb.remove(sheet1)

        self.column_head_color = PatternFill(start_color="24BBCD", end_color="24BBCD", fill_type="solid")
        self.cell_fill_color = PatternFill(start_color="B2D1D5", end_color="B2D1D5", fill_type="solid")

    def xl_file_write(self, data: list):
        """

        :param data:
        :return:
        """

        for ascii in range(ord("A"), ord("A")+ len(self.xl_column)+1):
            self.write_ws.column_dimensions[chr(ascii)].width = 12.75

        # 컬럼 셋업 ==============================================
        for i in range(len(self.xl_column)):
            cell = self.write_ws[f'{chr(ord("A")+i)}1'.format(i+1)]
            cell.value = self.xl_column[i]
            cell.alignment = Alignment(horizontal='center') # Text 중앙 정렬
            cell.font = Font(name="consolas")
            cell.fill = self.column_head_color
        # =======================================================

        # 데이터 셋업 =============================================
        for d in data:
            for i in range(len(self.xl_column)):
                # self.write_ws[f'{chr(ord("A") + i)}{d["num"] + 1}'.format(i + 1)] = d[self.xl_column[i]]
                cell = self.write_ws[f'{chr(ord("A") + i)}{d["num"] + 1}'.format(i + 1)]
                cell.value = d[self.xl_column[i]]

                if self.xl_column[i] in ["ROE", "5년ROE", "5년EPS성장률"]:
                    cell.number_format = '0.00"%"'
                elif self.xl_column[i] == "code":
                    cell.number_format = numbers.FORMAT_TEXT
                elif self.xl_column[i] in ["PER", "PBR", "EPS", "BPS"]:
                    cell.number_format = '0.000'

                cell.fill = self.cell_fill_color

        self.write_wb.save(self.reporting_file_path.format(_file_name_="주식"))
        # =======================================================
